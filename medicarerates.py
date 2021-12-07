from openpyxl import Workbook, load_workbook, worksheet

wb = load_workbook("medicare 2022 area 18.xlsx")
ws = wb.active

poslist = [i for i in range(1,100)]
cptlist = []

for x in range(10,11048 + 1):
    cptlist.append(ws[f"B{str(x)}"].value)

print("Please enter a CPT:")
cpt = input()

while cpt not in cptlist:
    print(f"Please enter a valid CPT!")
    cpt = input()

print("Please enter a Place of Service number:")
pos = input()

while int(pos) not in poslist:
    print(f"{pos} is an invalid POS. Please enter a value POS")
    pos = input()

print("Please enter the rate paid!")
rate = int(input())
rate = rate/100

def medicare2022rates(cpt,pos, rate):
    #Based on CPT and POS, returns rate 2022 Medicare rates area 18, Par Amount
    pos = int(pos)
    possiblevalues = float('inf')
    for row in range(10,11048 + 1):
        if ws[f"B{str(row)}"].value == cpt:
            price = ws[f"D{str(row)}"].value
            if pos != 22 and pos != 24:
                possiblevalues = price
                break
            if pos == 22 or pos == 24:
                if price < possiblevalues:
                    possiblevalues = price

    return f"CPT {cpt} rate is ${possiblevalues*rate:.2f}"

print(medicare2022rates(cpt,pos,rate))
